import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import json
import hashlib
import re
from datetime import datetime, date
from io import BytesIO
import base64

# ===================== CONFIGURAÇÃO DA PÁGINA =====================
st.set_page_config(
    page_title="Sistema PTM - Automação",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ===================== SISTEMA DE TEMAS PTM =====================
PALETAS = {
    "🌑 Dark Premium": {
        "primary": "#90f31e",
        "background": "#0c3b12",
        "secondaryBackground": "#158531",
        "textColor": "#eaf552",
    },
    "⚡ Neon Cyber": {
        "primary": "#00f5ff",
        "background": "#050510",
        "secondaryBackground": "#0d0d22",
        "textColor": "#e0f7ff",
    },
    "🟢 Matrix Green": {
        "primary": "#00ff41",
        "background": "#000d00",
        "secondaryBackground": "#001a00",
        "textColor": "#ccffcc",
    },
    "🔵 Hologram Blue": {
        "primary": "#4dc9ff",
        "background": "#020818",
        "secondaryBackground": "#061228",
        "textColor": "#d0eeff",
    },
    "🟣 Ultraviolet": {
        "primary": "#bf5fff",
        "background": "#0c0014",
        "secondaryBackground": "#160028",
        "textColor": "#eedeff",
    },
}

# ===================== SISTEMA DE ACESSO PTM =====================
_SEGREDO = "fyf9"
_DURACAO_MESES = 4
_ARQUIVO_CONFIG = "ptm_config_web.json"

def _hash(texto: str) -> str:
    return hashlib.sha256(texto.encode()).hexdigest()

def _carregar_config() -> dict:
    if os.path.exists(_ARQUIVO_CONFIG):
        try:
            with open(_ARQUIVO_CONFIG, 'r') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def _salvar_config(cfg: dict):
    with open(_ARQUIVO_CONFIG, 'w') as f:
        json.dump(cfg, f)

def _gerar_codigo_desbloqueio(id_usuario: str, ano: int, mes: int) -> str:
    base = f"UNLOCK{id_usuario.upper()},{mes:02d},{ano},{_SEGREDO}"
    h = hashlib.sha256(base.encode()).hexdigest()[:12].upper()
    return f"UNL-{h[:4]}-{h[4:8]}-{h[8:]}"

def _validar_codigo_desbloqueio(id_usuario: str, codigo: str):
    hoje = date.today()
    codigo = codigo.strip().upper()
    for delta in range(0, 7):
        mes_ref = hoje.month + delta
        ano_ref = hoje.year
        while mes_ref > 12:
            mes_ref -= 12
            ano_ref += 1
        esperado = _gerar_codigo_desbloqueio(id_usuario, ano_ref, mes_ref)
        if codigo == esperado:
            inicio = date(ano_ref, mes_ref, 1)
            fim_mes = inicio.month + _DURACAO_MESES
            fim_ano = inicio.year
            while fim_mes > 12:
                fim_mes -= 12
                fim_ano += 1
            fim = date(fim_ano, fim_mes, 1)
            return inicio, fim
    return None

def aplicar_tema(paleta_nome):
    """Aplica o tema selecionado"""
    paleta = PALETAS.get(paleta_nome, PALETAS["🌑 Dark Premium"])
    st.markdown(f"""
        <style>
        .stApp {{
            background-color: {paleta["background"]};
            color: {paleta["textColor"]};
        }}
        .stButton>button {{
            background-color: {paleta["primary"]};
            color: {paleta["background"]};
            border: none;
            border-radius: 5px;
            padding: 0.5rem 1rem;
            font-weight: bold;
        }}
        .stTextInput>div>div>input {{
            background-color: {paleta["secondaryBackground"]};
            color: {paleta["textColor"]};
        }}
        .stSelectbox>div>div>div {{
            background-color: {paleta["secondaryBackground"]};
            color: {paleta["textColor"]};
        }}
        .stTextArea>div>div>textarea {{
            background-color: {paleta["secondaryBackground"]};
            color: {paleta["textColor"]};
        }}
        h1, h2, h3 {{
            color: {paleta["primary"]};
        }}
        .success-box {{
            background-color: {paleta["secondaryBackground"]};
            padding: 1rem;
            border-radius: 5px;
            border-left: 5px solid {paleta["primary"]};
        }}
        </style>
    """, unsafe_allow_html=True)

def painel_autor_login():
    """Painel do autor para gerar códigos - versão para tela de login"""
    st.markdown("---")
    st.subheader("🔑 Painel do Autor — Gerador de Códigos")
    
    with st.expander("🔐 Acesso Restrito ao Autor", expanded=False):
        senha_master = st.text_input("Senha Master do Autor:", type="password", key="senha_master_login")
        
        if senha_master == "Johnwick10#":
            st.success("✅ Acesso autorizado!")
            
            col1, col2 = st.columns(2)
            
            with col1:
                id_usuario = st.text_input("ID do Usuário:", key="id_autor_login")
            
            with col2:
                hoje = date.today()
                meses = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                         "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
                mes_selecionado = st.selectbox("Mês de início:", 
                                               [f"{i+1:02d} - {m}" for i, m in enumerate(meses)],
                                               index=hoje.month - 1,
                                               key="mes_autor_login")
                ano_selecionado = st.selectbox("Ano de início:", 
                                              [str(hoje.year + i) for i in range(0, 4)],
                                              key="ano_autor_login")
            
            if st.button("⚙️ Gerar Código", key="btn_gerar_login"):
                if id_usuario:
                    mes = int(mes_selecionado.split(" - ")[0])
                    ano = int(ano_selecionado)
                    codigo = _gerar_codigo_desbloqueio(id_usuario, ano, mes)
                    
                    fim_mes = mes + _DURACAO_MESES
                    fim_ano = ano
                    while fim_mes > 12:
                        fim_mes -= 12
                        fim_ano += 1
                    
                    st.success(f"**Código gerado com sucesso!**")
                    st.code(codigo, language=None)
                    st.info(f"📅 Válido de 01/{mes:02d}/{ano} até 01/{fim_mes:02d}/{fim_ano}  |  Usuário: {id_usuario.upper()}")
                else:
                    st.warning("⚠️ Informe o ID do usuário!")
        elif senha_master:
            st.error("❌ Senha master incorreta!")

def verificar_acesso():
    """Verifica se o usuário tem acesso ao sistema"""
    if 'autenticado' not in st.session_state:
        st.session_state.autenticado = False
    
    if st.session_state.autenticado:
        return True
    
    cfg = _carregar_config()
    hoje = date.today()
    
    tem_cadastro = bool(cfg.get("senha_hash") and cfg.get("id_usuario"))
    acesso_ativo = False
    
    if tem_cadastro and cfg.get("validade_fim"):
        try:
            fim = date.fromisoformat(cfg["validade_fim"])
            inicio = date.fromisoformat(cfg.get("validade_inicio", str(hoje)))
            if inicio <= hoje <= fim:
                acesso_ativo = True
        except Exception:
            pass
    
    st.title("🔐 Sistema PTM — Autenticação")
    
    if not tem_cadastro:
        st.subheader("Primeiro Acesso — Configure seu Perfil")
        
        with st.form("form_cadastro"):
            id_usuario = st.text_input("Seu nome / ID de usuário:")
            senha1 = st.text_input("Crie sua senha de acesso:", type="password")
            senha2 = st.text_input("Confirme sua senha:", type="password")
            codigo = st.text_input("Código de desbloqueio (fornecido pelo autor):")
            
            col1, col2 = st.columns(2)
            with col1:
                submitted = st.form_submit_button("✅ Cadastrar e Entrar", type="primary", use_container_width=True)
            
            if submitted:
                if not id_usuario:
                    st.error("⚠️ Informe seu nome/ID!")
                elif len(senha1) < 4:
                    st.error("⚠️ Senha deve ter ao menos 4 caracteres!")
                elif senha1 != senha2:
                    st.error("❌ As senhas não coincidem!")
                elif not codigo:
                    st.error("⚠️ Informe o código de desbloqueio!")
                else:
                    res = _validar_codigo_desbloqueio(id_usuario, codigo)
                    if not res:
                        st.error("❌ Código de desbloqueio inválido!")
                    else:
                        inicio, fim = res
                        nova_cfg = {
                            "id_usuario": id_usuario.upper(),
                            "senha_hash": _hash(senha1),
                            "validade_inicio": str(inicio),
                            "validade_fim": str(fim),
                        }
                        _salvar_config(nova_cfg)
                        dias = (fim - hoje).days
                        st.success(f"✅ Bem-vindo, {id_usuario.upper()}!\n\n🔑 Acesso válido até: {fim.strftime('%d/%m/%Y')}\n📅 Dias restantes: {dias}")
                        st.session_state.autenticado = True
                        st.rerun()
        
        painel_autor_login()
    
    elif acesso_ativo:
        st.subheader(f"Olá, {cfg['id_usuario']}!")
        fim_dt = date.fromisoformat(cfg["validade_fim"])
        dias_restantes = (fim_dt - hoje).days
        st.info(f"🔑 Acesso válido até {fim_dt.strftime('%d/%m/%Y')} ({dias_restantes} dia(s))")
        
        with st.form("form_login"):
            senha = st.text_input("Digite sua senha:", type="password")
            
            col1, col2 = st.columns(2)
            with col1:
                submitted = st.form_submit_button("🔓 Entrar", type="primary", use_container_width=True)
            
            if submitted:
                if _hash(senha) == cfg["senha_hash"]:
                    st.session_state.autenticado = True
                    st.session_state.usuario = cfg['id_usuario']
                    st.success("✅ Login realizado com sucesso!")
                    st.rerun()
                else:
                    st.error("❌ Senha incorreta!")
        
        painel_autor_login()
    
    else:
        st.error("⌛ Seu acesso expirou!")
        if cfg.get("validade_fim"):
            fim_exp = date.fromisoformat(cfg["validade_fim"])
            st.warning(f"Expirou em: {fim_exp.strftime('%d/%m/%Y')}")
        
        st.info("Entre em contato com o autor para obter um novo código de desbloqueio.")
        
        with st.form("form_renovacao"):
            senha = st.text_input("Sua senha:", type="password")
            codigo = st.text_input("Novo código de desbloqueio:")
            
            col1, col2 = st.columns(2)
            with col1:
                submitted = st.form_submit_button("🔓 Renovar Acesso", type="primary", use_container_width=True)
            
            if submitted:
                if _hash(senha) != cfg.get("senha_hash", ""):
                    st.error("❌ Senha incorreta!")
                elif not codigo:
                    st.error("⚠️ Informe o código de desbloqueio!")
                else:
                    res = _validar_codigo_desbloqueio(cfg["id_usuario"], codigo)
                    if not res:
                        st.error("❌ Código de desbloqueio inválido!")
                    else:
                        inicio, fim = res
                        cfg["validade_inicio"] = str(inicio)
                        cfg["validade_fim"] = str(fim)
                        _salvar_config(cfg)
                        dias = (fim - hoje).days
                        st.success(f"✅ Acesso renovado!\n\n🔑 Válido até: {fim.strftime('%d/%m/%Y')}\n📅 Dias restantes: {dias}")
                        st.session_state.autenticado = True
                        st.rerun()
        
        painel_autor_login()
    
    return False

def painel_autor():
    """Painel do autor para gerar códigos - versão para quando está logado"""
    st.header("🔑 Painel do Autor — Gerador de Códigos")
    
    senha_master = st.text_input("Senha Master do Autor:", type="password", key="senha_master")
    
    if senha_master == "Johnwick10#":
        st.success("✅ Acesso autorizado!")
        
        col1, col2 = st.columns(2)
        
        with col1:
            id_usuario = st.text_input("ID do Usuário:")
        
        with col2:
            hoje = date.today()
            meses = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                     "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
            mes_selecionado = st.selectbox("Mês de início:", 
                                           [f"{i+1:02d} - {m}" for i, m in enumerate(meses)],
                                           index=hoje.month - 1)
            ano_selecionado = st.selectbox("Ano de início:", 
                                          [str(hoje.year + i) for i in range(0, 4)])
        
        if st.button("⚙️ Gerar Código", type="primary"):
            if id_usuario:
                mes = int(mes_selecionado.split(" - ")[0])
                ano = int(ano_selecionado)
                codigo = _gerar_codigo_desbloqueio(id_usuario, ano, mes)
                
                fim_mes = mes + _DURACAO_MESES
                fim_ano = ano
                while fim_mes > 12:
                    fim_mes -= 12
                    fim_ano += 1
                
                st.success(f"**Código gerado com sucesso!**")
                st.code(codigo, language=None)
                st.info(f"📅 Válido de 01/{mes:02d}/{ano} até 01/{fim_mes:02d}/{fim_ano}  |  Usuário: {id_usuario.upper()}")
            else:
                st.warning("⚠️ Informe o ID do usuário!")
    elif senha_master:
        st.error("❌ Senha master incorreta!")

def extrair_remessas(texto):
    """
    Extrai remessas de qualquer formato - VERSÃO CORRIGIDA
    Aceita: vírgulas, espaços, quebras de linha, tabulações
    """
    # Remove espaços extras e normaliza
    texto = texto.strip()
    
    # Substitui qualquer separador por espaço
    texto = re.sub(r'[,;\t\n\r]+', ' ', texto)
    
    # Extrai apenas números
    remessas = []
    for palavra in texto.split():
        palavra = palavra.strip()
        # Verifica se é um número válido
        if palavra.isdigit() and len(palavra) > 0:
            remessas.append(palavra)
    
    # Remove duplicatas mantendo a ordem
    remessas_unicas = []
    for r in remessas:
        if r not in remessas_unicas:
            remessas_unicas.append(r)
    
    return remessas_unicas

def gerar_planilha_custo(nome, dados, col_pedido, col_item, col_quantidade, arquivo_modelo):
    """Gera planilha CUSTO em memória"""
    try:
        wb = load_workbook(arquivo_modelo)
        ws = wb.active
        
        total_linhas = len(dados)
        percentual_por_linha = 1.0 / total_linhas
        
        linha = 7
        for _, row in dados.iterrows():
            ws[f'B{linha}'] = percentual_por_linha
            ws[f'B{linha}'].number_format = '0.00%'
            ws[f'O{linha}'] = row[col_pedido]
            ws[f'P{linha}'] = row[col_item]
            ws[f'Q{linha}'] = row[col_quantidade]
            linha += 1
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output, f"custo_{nome}.xlsx"
    except Exception as e:
        st.error(f"❌ Erro ao gerar {nome}: {str(e)}")
        return None, None

def gerar_planilha_cubagem(nome, dados, colunas, arquivo_espelho=None):
    """Gera planilha CUBAGEM em memória"""
    try:
        if arquivo_espelho:
            wb = load_workbook(arquivo_espelho)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cubagem"
            headers = {
                'B6': 'DESCRIÇÃO DO MATERIAL', 'C6': 'VALOR DO MATERIAL (R$)',
                'D6': 'TIPO DE EMBALAGEM', 'E6': 'QTDE', 'F6': 'COMP (m)',
                'G6': 'LARG (m)', 'H6': 'ALT (m)', 'I6': 'M³',
                'J6': 'PESO BRUTO UNIT(kg)', 'K6': 'PESO BRUTO TOTAL(kg)'
            }
            for cel, header in headers.items():
                ws[cel] = header
                ws[cel].font = Font(bold=True)
        
        linha = 7
        for _, row in dados.iterrows():
            ws[f'B{linha}'] = row[colunas[5]] if len(colunas) > 5 else ''
            ws[f'C{linha}'] = row[colunas[8]] if len(colunas) > 8 else ''
            ws[f'D{linha}'] = row[colunas[9]] if len(colunas) > 9 else ''
            ws[f'E{linha}'] = row[colunas[10]] if len(colunas) > 10 else ''
            ws[f'F{linha}'] = row[colunas[11]] if len(colunas) > 11 else ''
            ws[f'G{linha}'] = row[colunas[12]] if len(colunas) > 12 else ''
            ws[f'H{linha}'] = row[colunas[13]] if len(colunas) > 13 else ''
            ws[f'J{linha}'] = row[colunas[14]] if len(colunas) > 14 else ''
            linha += 1
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output, f"cubagem_{nome}.xlsx"
    except Exception as e:
        st.error(f"❌ Erro ao gerar {nome}: {str(e)}")
        return None, None

def modulo_custo():
    """Módulo de processamento CUSTO"""
    st.header("💰 Módulo CUSTO")
    
    tab1, tab2 = st.tabs(["🔧 Configuração", "🚀 Processamento"])
    
    with tab1:
        st.subheader("Configuração do Módulo CUSTO")
        
        arquivo_ptm = st.file_uploader("📂 Arquivo PTM (Aba PTM)", type=['xlsx', 'xlsm'], key="ptm_custo")
        arquivo_modelo = st.file_uploader("💰 Modelo CUSTO (Obrigatório)", type=['xlsx', 'xlsm'], key="modelo_custo")
        
        if arquivo_ptm and arquivo_modelo:
            st.success("✅ Arquivos carregados com sucesso!")
            st.session_state.arquivo_ptm_custo = arquivo_ptm
            st.session_state.arquivo_modelo_custo = arquivo_modelo
        else:
            st.warning("⚠️ Carregue os arquivos necessários para continuar")
    
    with tab2:
        st.subheader("Processamento CUSTO")
        
        if 'arquivo_ptm_custo' not in st.session_state or 'arquivo_modelo_custo' not in st.session_state:
            st.error("❌ Configure os arquivos na aba Configuração primeiro!")
            return
        
        st.info("💡 Digite as remessas (separadas por vírgula, espaço ou quebra de linha)")
        remessas_texto = st.text_area("🔢 Remessas:", height=150, key="remessas_custo", 
                                       placeholder="Exemplo:\n9202899470\n9202899479\n9202902088\n\nOu: 9202899470,9202899479,9202902088\n\nOu: 9202899470 9202899479 9202902088")
        
        if st.button("🚀 PROCESSAR CUSTO", type="primary"):
            if not remessas_texto:
                st.error("❌ Informe pelo menos uma remessa!")
                return
            
            with st.spinner("Processando..."):
                try:
                    # Carregar dados
                    dados_origem = pd.read_excel(st.session_state.arquivo_ptm_custo, sheet_name='PTM', engine='openpyxl')
                    
                    remessas = extrair_remessas(remessas_texto)
                    
                    if not remessas:
                        st.error("❌ Nenhuma remessa válida encontrada!")
                        return
                    
                    st.info(f"📊 Remessas detectadas ({len(remessas)}): {', '.join(remessas)}")
                    
                    colunas = list(dados_origem.columns)
                    col_remessa = colunas[0]
                    col_pedido = colunas[2]
                    col_item = colunas[3]
                    col_quantidade = colunas[6]
                    
                    remessas_float = [float(r) for r in remessas]
                    df_filtrado = dados_origem[dados_origem[col_remessa].isin(remessas_float)]
                    
                    if df_filtrado.empty:
                        st.warning("⚠️ Nenhum registro encontrado para as remessas informadas!")
                        return
                    
                    st.success(f"✅ Encontrados {len(df_filtrado)} registros")
                    
                    grupo_pedidos = df_filtrado.groupby(col_pedido)[col_remessa].nunique()
                    pedidos_multiplas = grupo_pedidos[grupo_pedidos > 1].index.tolist()
                    
                    if pedidos_multiplas:
                        st.warning(f"⚠️ Pedidos com múltiplas remessas: {len(pedidos_multiplas)}")
                    
                    # Salvar arquivo modelo temporariamente
                    with open("temp_modelo_custo.xlsx", "wb") as f:
                        f.write(st.session_state.arquivo_modelo_custo.getvalue())
                    
                    arquivos_gerados = []
                    
                    # Processar pedidos com múltiplas remessas
                    progress_bar = st.progress(0)
                    total_processamento = len(pedidos_multiplas) + len(remessas)
                    contador = 0
                    
                    for pedido in pedidos_multiplas:
                        dados = df_filtrado[df_filtrado[col_pedido] == pedido]
                        output, nome = gerar_planilha_custo(f"Pedido_{pedido}", dados, col_pedido, col_item, col_quantidade, "temp_modelo_custo.xlsx")
                        if output:
                            arquivos_gerados.append((output, nome))
                        contador += 1
                        progress_bar.progress(contador / total_processamento)
                    
                    # Processar remessas únicas
                    df_unicos = df_filtrado[~df_filtrado[col_pedido].isin(pedidos_multiplas)]
                    for remessa in remessas:
                        dados = df_unicos[df_unicos[col_remessa].astype(str).str.strip() == remessa]
                        if not dados.empty:
                            output, nome = gerar_planilha_custo(f"Remessa_{remessa}", dados, col_pedido, col_item, col_quantidade, "temp_modelo_custo.xlsx")
                            if output:
                                arquivos_gerados.append((output, nome))
                        contador += 1
                        progress_bar.progress(contador / total_processamento)
                    
                    progress_bar.empty()
                    
                    # Limpar arquivo temporário
                    if os.path.exists("temp_modelo_custo.xlsx"):
                        os.remove("temp_modelo_custo.xlsx")
                    
                    st.success(f"✅ {len(arquivos_gerados)} planilha(s) CUSTO gerada(s) com sucesso!")
                    
                    # Download dos arquivos
                    st.subheader("📥 Download dos Arquivos Gerados")
                    
                    col1, col2 = st.columns([3, 1])
                    with col1:
                        st.write(f"**Total de arquivos:** {len(arquivos_gerados)}")
                    
                    for idx, (output, nome) in enumerate(arquivos_gerados, 1):
                        col1, col2 = st.columns([4, 1])
                        with col1:
                            st.write(f"{idx}. {nome}")
                        with col2:
                            st.download_button(
                                label="⬇️ Baixar",
                                data=output,
                                file_name=nome,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_custo_{idx}"
                            )
                
                except Exception as e:
                    st.error(f"❌ Erro no processamento: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())

def modulo_cubagem():
    """Módulo de processamento CUBAGEM"""
    st.header("📊 Módulo CUBAGEM")
    
    tab1, tab2 = st.tabs(["🔧 Configuração", "🚀 Processamento"])
    
    with tab1:
        st.subheader("Configuração do Módulo CUBAGEM")
        
        arquivo_ptm = st.file_uploader("📂 Arquivo PTM (Aba PTM)", type=['xlsx', 'xlsm'], key="ptm_cubagem")
        arquivo_espelho = st.file_uploader("📋 Modelo/Espelho CUBAGEM (Opcional)", type=['xlsx', 'xlsm'], key="espelho_cubagem")
        
        if arquivo_ptm:
            st.success("✅ Arquivo PTM carregado!")
            st.session_state.arquivo_ptm_cubagem = arquivo_ptm
            if arquivo_espelho:
                st.session_state.arquivo_espelho_cubagem = arquivo_espelho
                st.success("✅ Modelo/Espelho carregado!")
        else:
            st.warning("⚠️ Carregue o arquivo PTM para continuar")
    
    with tab2:
        st.subheader("Processamento CUBAGEM")
        
        if 'arquivo_ptm_cubagem' not in st.session_state:
            st.error("❌ Configure o arquivo PTM na aba Configuração primeiro!")
            return
        
        st.info("💡 Digite as remessas (separadas por vírgula, espaço ou quebra de linha)")
        remessas_texto = st.text_area("🔢 Remessas:", height=150, key="remessas_cubagem",
                                       placeholder="Exemplo:\n9202899470\n9202899479\n9202902088\n\nOu: 9202899470,9202899479,9202902088\n\nOu: 9202899470 9202899479 9202902088")
        
        if st.button("🚀 PROCESSAR CUBAGEM", type="primary"):
            if not remessas_texto:
                st.error("❌ Informe pelo menos uma remessa!")
                return
            
            with st.spinner("Processando..."):
                try:
                    # Carregar dados
                    dados_origem = pd.read_excel(st.session_state.arquivo_ptm_cubagem, sheet_name='PTM', engine='openpyxl')
                    
                    remessas = extrair_remessas(remessas_texto)
                    
                    if not remessas:
                        st.error("❌ Nenhuma remessa válida encontrada!")
                        return
                    
                    st.info(f"📊 Remessas detectadas ({len(remessas)}): {', '.join(remessas)}")
                    
                    colunas = list(dados_origem.columns)
                    col_remessa = colunas[0]
                    col_pedido = colunas[2]
                    
                    remessas_float = [float(r) for r in remessas]
                    df_filtrado = dados_origem[dados_origem[col_remessa].isin(remessas_float)]
                    
                    if df_filtrado.empty:
                        st.warning("⚠️ Nenhum registro encontrado para as remessas informadas!")
                        return
                    
                    st.success(f"✅ Encontrados {len(df_filtrado)} registros")
                    
                    grupo_pedidos = df_filtrado.groupby(col_pedido)[col_remessa].nunique()
                    pedidos_repetidos = grupo_pedidos[grupo_pedidos > 1].index.tolist()
                    
                    if pedidos_repetidos:
                        st.warning(f"⚠️ Pedidos repetidos: {len(pedidos_repetidos)}")
                    
                    # Salvar arquivo espelho temporariamente se existir
                    arquivo_espelho_temp = None
                    if 'arquivo_espelho_cubagem' in st.session_state:
                        with open("temp_espelho_cubagem.xlsx", "wb") as f:
                            f.write(st.session_state.arquivo_espelho_cubagem.getvalue())
                        arquivo_espelho_temp = "temp_espelho_cubagem.xlsx"
                    
                    arquivos_gerados = []
                    
                    # Processar pedidos repetidos
                    progress_bar = st.progress(0)
                    total_processamento = len(pedidos_repetidos) + len(remessas)
                    contador = 0
                    
                    for pedido in pedidos_repetidos:
                        dados = df_filtrado[df_filtrado[col_pedido] == pedido]
                        output, nome = gerar_planilha_cubagem(f"Pedido_{pedido}", dados, colunas, arquivo_espelho_temp)
                        if output:
                            arquivos_gerados.append((output, nome))
                        contador += 1
                        progress_bar.progress(contador / total_processamento)
                    
                    # Processar remessas únicas
                    df_unicos = df_filtrado[~df_filtrado[col_pedido].isin(pedidos_repetidos)]
                    for remessa in remessas:
                        dados = df_unicos[df_unicos[col_remessa].astype(str).str.strip() == remessa]
                        if not dados.empty:
                            output, nome = gerar_planilha_cubagem(remessa, dados, colunas, arquivo_espelho_temp)
                            if output:
                                arquivos_gerados.append((output, nome))
                        contador += 1
                        progress_bar.progress(contador / total_processamento)
                    
                    progress_bar.empty()
                    
                    # Limpar arquivo temporário
                    if arquivo_espelho_temp and os.path.exists(arquivo_espelho_temp):
                        os.remove(arquivo_espelho_temp)
                    
                    st.success(f"✅ {len(arquivos_gerados)} planilha(s) CUBAGEM gerada(s) com sucesso!")
                    
                    # Download dos arquivos
                    st.subheader("📥 Download dos Arquivos Gerados")
                    
                    col1, col2 = st.columns([3, 1])
                    with col1:
                        st.write(f"**Total de arquivos:** {len(arquivos_gerados)}")
                    
                    for idx, (output, nome) in enumerate(arquivos_gerados, 1):
                        col1, col2 = st.columns([4, 1])
                        with col1:
                            st.write(f"{idx}. {nome}")
                        with col2:
                            st.download_button(
                                label="⬇️ Baixar",
                                data=output,
                                file_name=nome,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_cubagem_{idx}"
                            )
                
                except Exception as e:
                    st.error(f"❌ Erro no processamento: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())

def main():
    """Função principal"""
    
    # Carregar configurações
    cfg = _carregar_config()
    tema_selecionado = cfg.get("tema", "🌑 Dark Premium")
    
    # Aplicar tema
    aplicar_tema(tema_selecionado)
    
    # Verificar acesso
    if not verificar_acesso():
        return
    
    # Sidebar
    with st.sidebar:
        st.title("🏭 Sistema PTM")
        st.markdown("---")
        
        # Seleção de tema
        novo_tema = st.selectbox("🎨 Tema:", list(PALETAS.keys()), index=list(PALETAS.keys()).index(tema_selecionado))
        if novo_tema != tema_selecionado:
            cfg["tema"] = novo_tema
            _salvar_config(cfg)
            st.rerun()
        
        st.markdown("---")
        
        # Informações do usuário
        if 'usuario' in st.session_state:
            st.info(f"👤 Usuário: {st.session_state.usuario}")
        
        # Botão de logout
        if st.button("🚪 Sair", use_container_width=True):
            st.session_state.autenticado = False
            st.rerun()
        
        st.markdown("---")
        st.caption("Desenvolvido por DJALMA Barbosa (FYF9) e Thiago Aniz (F09S) • 2026")
    
    # Conteúdo principal
    st.title("🏭 SISTEMA DE AUTOMAÇÃO PTM")
    st.subheader("Automatizador Inteligente de CUSTO e CUBAGEM")
    
    # Tabs principais
    tab1, tab2, tab3, tab4 = st.tabs(["🏠 HOME", "💰 CUSTO", "📊 CUBAGEM", "🔑 PAINEL DO AUTOR"])
    
    with tab1:
        st.header("Bem-vindo ao Sistema PTM")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.info("""
            ### 💰 Módulo CUSTO
            - Processamento automático de planilhas de custo
            - Distribuição percentual automática
            - Suporte a múltiplas remessas
            - Processamento em lote
            """)
        
        with col2:
            st.info("""
            ### 📊 Módulo CUBAGEM
            - Geração automática de planilhas de cubagem
            - Suporte a modelo/espelho personalizado
            - Processamento em lote
            - Download individual de arquivos
            """)
        
        st.success("""
        ### 🚀 Como usar:
        1. Selecione o módulo desejado (CUSTO ou CUBAGEM)
        2. Configure os arquivos necessários
        3. Digite as remessas a processar (aceita qualquer formato)
        4. Clique em processar e baixe os arquivos gerados
        
        ### 📋 Formatos aceitos para remessas:
        - **Uma por linha:** 9202899470, 9202899479, 9202902088
        - **Separadas por vírgula:** 9202899470,9202899479,9202902088
        - **Separadas por espaço:** 9202899470 9202899479 9202902088
        - **Formato misto:** 9202899470, 9202899479 9202902088
        """)
    
    with tab2:
        modulo_custo()
    
    with tab3:
        modulo_cubagem()
    
    with tab4:
        painel_autor()

if __name__ == "__main__":
    main()
